#!/usr/bin/env python3
"""Generate the Delta Crown owner-decision Excel workbook.

Run with:
    uv run --with openpyxl python tools/generate_owner_decision_workbook.py

Why openpyxl?
    Excel is picky. Hand-rolled XLSX XML is a crime scene waiting to happen.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path("generated/delta-crown-owner-decision-workbook.xlsx")

TEAL = "0B3D3D"
GOLD = "C9A227"
LIGHT_GOLD = "F5E6A8"
LIGHT_BLUE = "DDEBF7"
LIGHT_GREEN = "E2F0D9"
LIGHT_YELLOW = "FFF2CC"
LIGHT_RED = "FCE4D6"
WHITE = "FFFFFF"
DARK = "1F1F1F"
BORDER = "D9E2F3"


@dataclass(frozen=True)
class Option:
    key: str
    description: str
    pros: str
    cons: str


@dataclass(frozen=True)
class Decision:
    decision_id: str
    title: str
    current_state: str
    risk: str
    recommendation: str
    owner_needed: str
    evidence: tuple[str, ...]
    options: tuple[Option, ...]
    followups: tuple[str, ...]


DECISIONS: tuple[Decision, ...] = (
    Decision(
        "OD-001",
        "GitHub default branch / stale main",
        "Remote default branch is main; canonical current project work and GitHub Pages source are on gh-pages.",
        "Future operators may land on stale architecture and treat it as current.",
        "Prefer changing GitHub default branch to gh-pages. If that is not possible, make main a hard redirect/archive branch.",
        "Tyler / repo owner",
        (
            "git remote show origin reports HEAD branch: main.",
            "README says gh-pages is canonical and main is abandoned earlier architecture.",
            "Live public site publishes from gh-pages.",
        ),
        (
            Option("A", "Change GitHub default branch to gh-pages.", "Repo landing page matches current work.", "Requires GitHub repo admin action."),
            Option("B", "Keep main but replace README with a clear redirect to gh-pages.", "Fast and low risk.", "Users can still browse stale files."),
            Option("C", "Archive/delete stale main after preservation tag.", "Cleanest long-term.", "Higher governance risk; needs explicit approval."),
        ),
        (
            "If A: update GitHub repository default branch setting.",
            "If B: patch main README to point to gh-pages and warn it is legacy.",
            "If C: tag legacy main, verify no workflows depend on it, then archive/delete.",
        ),
    ),
    Decision(
        "OD-002",
        "Brand Resources vs Brand Assets model",
        "Docs target Brand Resources; tenant/docs still include Marketing Brand Assets and legacy ClientServices artifacts.",
        "Users and MSP operators may confuse reference material, marketing execution assets, legacy ClientServices, and client records.",
        "Use Brand Resources as the staff-facing reference concept; keep Brand Assets as Marketing execution/library if needed. Do not repurpose ClientServices yet.",
        "Tyler + business/content owner",
        (
            "docs/brand-resources-target-model.md defines Brand Resources as reusable reference material, not CRM/client records.",
            "docs/clientservices-to-brand-resources-transition-plan.md says ClientServices assumptions are legacy.",
            "docs/legacy-clientservices-cleanup-register.md prohibits tenant cleanup without inventory and approval.",
        ),
        (
            Option("A", "Brand Resources = staff reference; Brand Assets = Marketing execution library.", "Clear separation of audience and purpose.", "Requires explaining two related terms."),
            Option("B", "Consolidate everything to Brand Resources.", "Simplest user-facing label.", "Unsafe before dependency/content review."),
            Option("C", "Keep Brand Assets as the only user-facing term.", "Matches current Marketing wording.", "Weakens replacement of ClientServices assumptions."),
            Option("D", "Defer until owners classify current usage.", "Avoids premature naming decision.", "Slows launch cleanup."),
        ),
        (
            "Approve final vocabulary.",
            "Create follow-up Brand Resources implementation issue after decision.",
            "Do not rename/delete/repurpose live ClientServices resources yet.",
        ),
    ),
    Decision(
        "OD-003",
        "Dynamic security group owners",
        "AllStaff, Managers, Marketing, Stylists, and External have zero owners in Graph.",
        "Nobody is clearly accountable when metadata, membership, or access drifts.",
        "Assign at least two owners per group: MSP/operator plus business/access owner.",
        "Tyler / Megan",
        (
            "docs/delta-crown-identity-inventory-summary.md shows zero owners for all five dynamic groups.",
            "Current group counts: AllStaff=6, Managers=1, Marketing=0, Stylists=0, External=0.",
            "Production launch readiness report marks named operational owners as a no-go item.",
        ),
        (
            Option("A", "Two owners per group: MSP operator + business/access owner.", "Best production governance.", "Requires naming humans now."),
            Option("B", "One central identity owner for all dynamic groups.", "Fastest owner assignment.", "Weak business accountability."),
            Option("C", "Keep ownerless until metadata cleanup finishes.", "Avoids churn.", "Leaves governance gap open."),
        ),
        (
            "Name owner candidates for each group.",
            "Create tenant change issue for owner assignment.",
            "Verify owner assignment via Graph after change.",
        ),
    ),
    Decision(
        "OD-004",
        "DLP test mode vs enforce path",
        "DCE-Data-Protection and Corp-Data-Protection are TestWithNotifications; External-Sharing-Block is enabled.",
        "Test mode is not final enforcement. Claims about production DLP must distinguish test from enforce.",
        "Review raw DLP rule details, then choose a staged enforce flip unless rule review proves very low risk.",
        "Tyler + security/compliance owner",
        (
            "docs/delta-crown-compliance-inventory-summary.md lists 6 DLP policies and 8 rules.",
            "DCE-Data-Protection mode: TestWithNotifications.",
            "Corp-Data-Protection mode: TestWithNotifications.",
            "External-Sharing-Block mode: Enable.",
        ),
        (
            Option("A", "Keep test mode for 30 days and review alerts.", "Lower business disruption risk.", "Not fully enforcing."),
            Option("B", "Flip both policies to enforce in one window.", "Fastest enforcement posture.", "Higher false-positive risk."),
            Option("C", "Enforce one policy first, then the other.", "Balanced staged rollout.", "Takes longer."),
            Option("D", "Keep test mode indefinitely.", "Avoids disruption.", "Not production-grade."),
        ),
        (
            "Review local raw DLP rule exports.",
            "Pick enforce order and maintenance window.",
            "Create change issue with rollback owner and validation plan.",
        ),
    ),
    Decision(
        "OD-005",
        "DeltaCrown-TeamsProvisioner-TEMP app fate",
        "TEMP app registration exists with three password credentials expired on 2026-04-16.",
        "Expired temporary apps are security/governance clutter unless dependency and owner are documented.",
        "Delete the app if no dependency remains. Otherwise rename/document owner and remove expired secrets.",
        "Tyler / tenant app owner",
        (
            "docs/delta-crown-security-apps-licenses-inventory-summary.md lists DeltaCrown-TeamsProvisioner-TEMP.",
            "Inventory found three expired app password credentials for the TEMP app.",
            "Safety notes say not to delete app registrations without dependency review and approval.",
        ),
        (
            Option("A", "Delete app after dependency check.", "Cleanest if no longer used.", "Requires confidence no automation depends on it."),
            Option("B", "Keep, rename, assign owner, remove expired credentials, document purpose.", "Safe if still needed.", "Keeps extra app surface."),
            Option("C", "Keep as-is temporarily with deadline.", "Buys time.", "Still leaves known smell."),
        ),
        (
            "Confirm whether any Teams provisioning automation still uses the app.",
            "Create cleanup issue for delete or rename/credential cleanup.",
            "Verify app inventory after cleanup.",
        ),
    ),
    Decision(
        "OD-006",
        "ClientServices deprecation banner / cleanup",
        "Legacy ClientServices artifacts are documented as empty/broad and client records are out of M365 scope.",
        "Users may treat old ClientServices resources as approved client-data storage, conflicting with current scope.",
        "Add a visible deprecation/owner note first. Do not delete, rename, or repurpose until owner approval and cleanup plan exist.",
        "Tyler + data/content owner",
        (
            "docs/legacy-clientservices-cleanup-register.md states ClientServices/client-record assumptions are legacy.",
            "docs/delta-crown-sharepoint-pnp-inventory-summary.md documents ClientServices artifacts and permissions metadata.",
            "Brand Resources target model says client records are out of Microsoft 365 scope.",
        ),
        (
            Option("A", "Add deprecation/banner notice: do not use for client records; pending cleanup.", "Immediate clear signal without destructive changes.", "Requires tenant page/banner edit."),
            Option("B", "Remove navigation links only.", "Reduces accidental use.", "Does not explain what happened."),
            Option("C", "Archive or delete after owner approval.", "Cleaner long-term.", "Unsafe before dependency/content approval."),
            Option("D", "Repurpose as Brand Resources after approval.", "May reuse existing resource.", "Risky URL/history/permission baggage."),
        ),
        (
            "Approve banner wording.",
            "Create tenant change issue for non-destructive deprecation signal.",
            "Decide later: archive, replace, repurpose, or remove.",
        ),
    ),
)


class Formats:
    def __init__(self) -> None:
        thin = Side(style="thin", color=BORDER)
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.title = Font(name="Aptos Display", size=18, bold=True, color=TEAL)
        self.subtitle = Font(name="Aptos", size=11, italic=True, color="666666")
        self.header = Font(name="Aptos", size=11, bold=True, color=WHITE)
        self.section = Font(name="Aptos", size=13, bold=True, color=TEAL)
        self.body = Font(name="Aptos", size=11, color=DARK)
        self.body_bold = Font(name="Aptos", size=11, bold=True, color=DARK)
        self.white_bold = Font(name="Aptos", size=11, bold=True, color=WHITE)
        self.fill_header = PatternFill("solid", fgColor=TEAL)
        self.fill_gold = PatternFill("solid", fgColor=LIGHT_GOLD)
        self.fill_blue = PatternFill("solid", fgColor=LIGHT_BLUE)
        self.fill_green = PatternFill("solid", fgColor=LIGHT_GREEN)
        self.fill_yellow = PatternFill("solid", fgColor=LIGHT_YELLOW)
        self.fill_red = PatternFill("solid", fgColor=LIGHT_RED)
        self.center = Alignment(horizontal="center", vertical="top", wrap_text=True)
        self.top = Alignment(vertical="top", wrap_text=True)


def write_title(ws, title: str, subtitle: str, fmt: Formats, width: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.cell(1, 1, title).font = fmt.title
    ws.cell(1, 1).alignment = fmt.top
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    ws.cell(2, 1, subtitle).font = fmt.subtitle
    ws.cell(2, 1).alignment = fmt.top
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 34
    return 4


def write_table(ws, start_row: int, headers: list[str], rows: Iterable[Iterable[object]], fmt: Formats, table_name: str | None = None) -> int:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = fmt.header
        cell.fill = fmt.fill_header
        cell.alignment = fmt.center
        cell.border = fmt.border
    row_num = start_row + 1
    for row in rows:
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_num, col, value)
            cell.font = fmt.body
            cell.alignment = fmt.top
            cell.border = fmt.border
        row_num += 1
    if table_name and row_num > start_row + 1:
        end_col = get_column_letter(len(headers))
        ref = f"A{start_row}:{end_col}{row_num - 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    return row_num + 1


def set_widths(ws, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def freeze_and_filter(ws, row: int = 5) -> None:
    ws.freeze_panes = f"A{row}"


def style_status_cells(ws, col: int, fmt: Formats) -> None:
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, col)
        value = str(cell.value or "").lower()
        if "not started" in value or "tbd" in value:
            cell.fill = fmt.fill_yellow
        elif "blocked" in value or "no-go" in value:
            cell.fill = fmt.fill_red
        elif "pass" in value or "go" in value:
            cell.fill = fmt.fill_green


def dashboard(wb: Workbook, fmt: Formats) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    row = write_title(ws, "Delta Crown Owner Decision Workbook", "Launch-blocking owner decisions, recommendations, evidence, and follow-up actions", fmt, 5)
    data = [
        ("Workbook purpose", "Owner launch decisions", "Converts launch blockers into explicit choices."),
        ("Decision count", len(DECISIONS), "Six decisions currently block full production-launch claims."),
        ("Recommended launch posture", "Controlled pilot / owner validation", "Built and hardened, but not full production launch yet."),
        ("Top technical blocker", "Teams channel read context", "Members readable; team/channel detail still blocked."),
        ("Top governance blocker", "Named owners + DLP posture", "Needs accountable owners and enforcement decisions."),
    ]
    row = write_table(ws, row, ["Metric", "Value", "Why it matters"], data, fmt, "DashboardTable")
    ws.cell(row, 1, "How to use this workbook").font = fmt.section
    instructions = [
        "1. Start on Decision Matrix and pick a Tyler decision for each row.",
        "2. Use OD-001 through OD-006 tabs for evidence, risks, options, and recommendation detail.",
        "3. Use Action Tracker to split approved work into small follow-up change issues.",
        "4. Do not make tenant-impacting changes from this workbook alone; create tracked change issues first.",
    ]
    for offset, text in enumerate(instructions, start=1):
        ws.cell(row + offset, 1, text).alignment = fmt.top
    set_widths(ws, [26, 34, 72, 20, 20])
    freeze_and_filter(ws)


def decision_matrix(wb: Workbook, fmt: Formats) -> None:
    ws = wb.create_sheet("Decision Matrix")
    row = write_title(ws, "Decision Matrix", "One-line view of what Tyler/Megan need to decide", fmt, 9)
    rows = [
        [d.decision_id, d.title, d.current_state, d.risk, d.recommendation, d.owner_needed, "TBD", "TBD", ""]
        for d in DECISIONS
    ]
    write_table(ws, row, ["ID", "Decision", "Current state", "Risk", "Recommendation", "Owner needed", "Tyler decision", "Decision date", "Notes"], rows, fmt, "DecisionMatrix")
    set_widths(ws, [11, 30, 46, 44, 50, 24, 22, 16, 34])
    style_status_cells(ws, 7, fmt)
    freeze_and_filter(ws)


def action_tracker(wb: Workbook, fmt: Formats) -> None:
    ws = wb.create_sheet("Action Tracker")
    row = write_title(ws, "Action Tracker", "Follow-up actions after decisions are approved", fmt, 7)
    rows = []
    for d in DECISIONS:
        for action in d.followups:
            rows.append([d.decision_id, action, "TBD", d.owner_needed, "Not started", "TBD", ""])
    write_table(ws, row, ["Decision ID", "Follow-up action", "Type", "Owner", "Status", "Target date", "Notes"], rows, fmt, "ActionTracker")
    set_widths(ws, [13, 62, 18, 28, 18, 16, 36])
    style_status_cells(ws, 5, fmt)
    freeze_and_filter(ws)


def evidence_sheet(wb: Workbook, fmt: Formats) -> None:
    ws = wb.create_sheet("Evidence")
    row = write_title(ws, "Evidence", "Supporting evidence behind recommendations", fmt, 3)
    rows = []
    for d in DECISIONS:
        for evidence in d.evidence:
            rows.append([d.decision_id, evidence, "See repo docs and owner-decision worksheet."])
    write_table(ws, row, ["Decision ID", "Evidence", "Source / context"], rows, fmt, "EvidenceTable")
    set_widths(ws, [14, 88, 46])
    freeze_and_filter(ws)


def glossary(wb: Workbook, fmt: Formats) -> None:
    ws = wb.create_sheet("Glossary")
    row = write_title(ws, "Glossary", "Plain-English definitions for decision terms", fmt, 2)
    rows = [
        ("DDG", "Dynamic Distribution Group for Exchange mail routing."),
        ("Dynamic security group", "Entra group whose membership is calculated from user attributes."),
        ("TestWithNotifications", "DLP policy mode that warns/tests without full enforcement."),
        ("Brand Resources", "Approved staff-facing reference material, not client records."),
        ("Brand Assets", "Marketing execution/asset library concept."),
        ("ClientServices", "Legacy concept/artifacts; not approved client-record storage in M365."),
        ("Controlled pilot", "Limited validation state before full launch approval."),
    ]
    write_table(ws, row, ["Term", "Meaning"], rows, fmt, "GlossaryTable")
    set_widths(ws, [26, 90])
    freeze_and_filter(ws)


def decision_detail(wb: Workbook, fmt: Formats, decision: Decision) -> None:
    ws = wb.create_sheet(decision.decision_id)
    row = write_title(ws, f"{decision.decision_id} — {decision.title}", "Decision detail, options, evidence, and action prompts", fmt, 4)
    overview = [
        ("Current state", decision.current_state),
        ("Risk", decision.risk),
        ("Recommendation", decision.recommendation),
        ("Owner needed", decision.owner_needed),
        ("Tyler selected option", "TBD"),
        ("Approval / date", "TBD"),
        ("Follow-up issue", "TBD"),
    ]
    row = write_table(ws, row, ["Field", "Detail"], overview, fmt, f"{decision.decision_id.replace('-', '')}Overview")
    ws.cell(row, 1, "Options").font = fmt.section
    row += 1
    row = write_table(
        ws,
        row,
        ["Option", "Description", "Pros", "Cons"],
        [[o.key, o.description, o.pros, o.cons] for o in decision.options],
        fmt,
        f"{decision.decision_id.replace('-', '')}Options",
    )
    ws.cell(row, 1, "Supporting evidence").font = fmt.section
    row += 1
    row = write_table(ws, row, ["Evidence", "Detail"], [["Evidence", item] for item in decision.evidence], fmt, None)
    ws.cell(row, 1, "Follow-up actions").font = fmt.section
    row += 1
    write_table(ws, row, ["Action", "Detail"], [["Action", action] for action in decision.followups], fmt, None)
    set_widths(ws, [24, 70, 50, 50])
    freeze_and_filter(ws)


def apply_global_style(wb: Workbook, fmt: Formats) -> None:
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment
        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 36


def build_workbook() -> Workbook:
    wb = Workbook()
    fmt = Formats()
    dashboard(wb, fmt)
    decision_matrix(wb, fmt)
    action_tracker(wb, fmt)
    evidence_sheet(wb, fmt)
    glossary(wb, fmt)
    for decision in DECISIONS:
        decision_detail(wb, fmt, decision)
    apply_global_style(wb, fmt)
    return wb


def validate_workbook(path: Path) -> None:
    loaded = load_workbook(path)
    expected = {"Dashboard", "Decision Matrix", "Action Tracker", "Evidence", "Glossary"} | {d.decision_id for d in DECISIONS}
    actual = set(loaded.sheetnames)
    missing = expected - actual
    if missing:
        raise RuntimeError(f"Workbook missing sheets: {sorted(missing)}")
    if loaded["Decision Matrix"].max_row < len(DECISIONS) + 4:
        raise RuntimeError("Decision Matrix does not contain expected decision rows")


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUT)
    validate_workbook(OUT)
    print(f"Wrote valid workbook: {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
